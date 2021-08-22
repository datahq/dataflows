import datetime
from dataflows.base.exceptions import ProcessorError
import pytest
from dataflows import Flow


data = [
    dict(x=1, y='a'),
    dict(x=2, y='b'),
    dict(x=3, y='c'),
]


def test_dump_to_sql():
    from dataflows import Flow, printer, dump_to_sql
    from sqlalchemy import create_engine

    f = Flow(
        data,
        printer(),
        dump_to_sql(dict(
                output_table={
                    'resource-name': 'res_1'
                }
            ),
            engine='sqlite:///out/test.db')
    )
    f.process()

    # Check validity
    engine = create_engine('sqlite:///out/test.db')
    result = list(dict(x) for x in engine.execute('select * from output_table'))
    assert result == data


def test_add_computed_field():
    from dataflows import add_computed_field
    f = Flow(
        data,
        add_computed_field([
            dict(source=['x', 'x'], target='xx', operation='multiply'),
            dict(target='f', operation='format', with_='{y} - {x}')
        ])
    )
    results, dp, stats = f.results()
    results = list(results[0])

    xx = [x['xx'] for x in results]
    f = [x['f'] for x in results]

    assert xx == [1, 4, 9]
    assert f == ['a - 1', 'b - 2', 'c - 3']


def test_add_computed_field_func():
    from dataflows import add_computed_field

    data = [
        dict(x=i) for i in range(3)
    ]

    f = Flow(
        data,
        add_computed_field([
            dict(target=dict(name='sq', type='integer'),
                 operation=lambda row: row['x'] ** 2),
            dict(target='f', operation='format', with_='{x} - {x}')
        ])
    )
    results, *_ = f.results()
    results = list(results[0])

    assert results == [
        dict(x=0, sq=0, f='0 - 0'),
        dict(x=1, sq=1, f='1 - 1'),
        dict(x=2, sq=4, f='2 - 2'),
    ]


def test_add_metadata():
    from dataflows import add_metadata
    f = Flow(
        data,
        add_metadata(author='Adam Kariv')
    )
    _, dp, _ = f.results()
    assert dp.descriptor['author'] == 'Adam Kariv'


def test_delete_field():
    from dataflows import delete_fields
    data2 = [
        dict(x=1, y='a', xx=True),
        dict(x=2, y='b', xx=True),
        dict(x=3, y='c', xx=True),
    ]
    f = Flow(
        data,
        delete_fields(['x'])
    )
    results, dp, _ = f.results()
    for i in results[0]:
        assert list(i.keys()) == ['y']
    assert dp.descriptor['resources'][0]['schema']['fields'] == \
        [dict(name='y', type='string', format='default')]

    f = Flow(
        data,
        data2,
        delete_fields(['x+'])
    )
    results, dp, _ = f.results()
    for res in results:
        for i in res:
            assert list(i.keys()) == ['y']
    assert dp.descriptor['resources'][0]['schema']['fields'] == \
        [dict(name='y', type='string', format='default')]


def test_select_field_no_regex():
    from dataflows import select_fields
    data = [
        {'x':i, '.':i}
        for i in range(3)
    ]

    f = Flow(
        data,
        select_fields(['.'], regex=False)
    )
    results, dp, _ = f.results()
    for i in results[0]:
        assert list(i.keys()) == ['.']
    assert dp.descriptor['resources'][0]['schema']['fields'] == \
        [dict(name='.', type='integer', format='default')]


def test_select_field_regex():
    from dataflows import select_fields
    data = [
        dict(x=i, y2=i, y1=i)
        for i in range(3)
    ]

    f = Flow(
        data,
        select_fields(['y\d'])
    )
    results, dp, _ = f.results()
    for i in results[0]:
        assert list(i.keys()) == ['y2', 'y1']
    assert dp.descriptor['resources'][0]['schema']['fields'] == \
        [
            dict(name='y2', type='integer', format='default'),
            dict(name='y1', type='integer', format='default')
        ]

def test_select_field_no_fields_selected():
    from dataflows import select_fields, exceptions
    data = [
        dict(x=i, y2=i, y1=i)
        for i in range(3)
    ]

    f = Flow(
        data,
        select_fields(['x\d'])
    )
    with pytest.raises(exceptions.ProcessorError):
        results, dp, _ = f.results()
        print(results)
        print(dp)


def test_find_replace():
    from dataflows import find_replace
    f = Flow(
        data,
        find_replace([dict(
            name='y',
            patterns=[
                dict(find='a', replace='Apple'),
                dict(find='b', replace='Banana'),
                dict(find='c', replace='Coconut'),
            ]
        )])
    )
    results, _, _ = f.results()
    y = [r['y'] for r in results[0]]
    assert y == ['Apple', 'Banana', 'Coconut']


def test_unpivot_no_regex():
    from dataflows import unpivot
    data = [
        dict([('[.]', i), ('[,+', str(i))]) for i in range(3)
    ]
    f = Flow(
        data,
        unpivot(
            [
                dict(
                    name='[.]',
                    keys=dict(
                        field='x'
                    )
                ),
                dict(
                    name='[,+',
                    keys=dict(
                        field='y'
                    )
                ),
            ],
            [
                dict(
                    name='field',
                    type='string'
                )
            ],
            dict(
                name='the-value',
                type='any'
            ), regex=False
        )
    )
    results, _, _ = f.results()
    assert results[0] == [
        dict(zip(['field', 'the-value'], r))
        for r in
        [
            ['x', 0],
            ['y', '0'],
            ['x', 1],
            ['y', '1'],
            ['x', 2],
            ['y', '2'],
        ]
    ]


def test_unpivot():
    from dataflows import unpivot
    f = Flow(
        data,
        unpivot(
            [
                dict(
                    name='x',
                    keys=dict(
                        field='x-value'
                    )
                ),
                dict(
                    name='y',
                    keys=dict(
                        field='y-value'
                    )
                ),
            ],
            [
                dict(
                    name='field',
                    type='string'
                )
            ],
            dict(
                name='the-value',
                type='any'
            )
        )
    )
    results, _, _ = f.results()
    assert results[0] == [
        dict(zip(['field', 'the-value'], r))
        for r in
        [
            ['x-value', 1],
            ['y-value', 'a'],
            ['x-value', 2],
            ['y-value', 'b'],
            ['x-value', 3],
            ['y-value', 'c'],
        ]
    ]

def test_unpivot_simple():
    from dataflows import unpivot
    up_data = [
        dict(k='a', v1=1, v2=2, v3=3),
        dict(k='b', v1=2, v2=4, v3=6),
    ]

    f = Flow(
        up_data,
        unpivot(
            [
                dict(
                    name=r'v\d',
                    keys=dict()
                ),
            ],
            [],
            dict(
                name='v',
                type='integer'
            )
        )
    )
    results, _, _ = f.results()
    assert results[0] == [
        dict(zip(['k', 'v'], r))
        for r in
        [
            ['a', 1],
            ['a', 2],
            ['a', 3],
            ['b', 2],
            ['b', 4],
            ['b', 6],
        ]
    ]


def test_unpivot_any_resources():
    from dataflows import unpivot, validate
    data1 = [
        dict(
            [('name', 'ike{}'.format(i))] +
            [(str(year), year + i) for year in range(1990, 2020, 10)]
        )
        for i in range(5)
    ]
    data2 = [
        dict(
            [('city', 'mike{}'.format(i))] +
            [(str(year), year + i) for year in range(2050, 2080, 10)]
        )
        for i in range(5)
    ]
    f = Flow(
        data1,
        data2,
        unpivot(
            [
                dict(
                    name='([0-9]+)',
                    keys=dict(
                        year='\\1'
                    )
                )
            ],
            [
                dict(
                    name='year',
                    type='integer'
                )
            ],
            dict(
                name='amount',
                type='integer'
            )
        ),
        validate()
    )
    results, _, _ = f.results()
    assert results[0] == [
        dict(zip(['name', 'year', 'amount'], r))
        for r in
        [
            ['ike0', 1990, 1990],
            ['ike0', 2000, 2000],
            ['ike0', 2010, 2010],
            ['ike1', 1990, 1991],
            ['ike1', 2000, 2001],
            ['ike1', 2010, 2011],
            ['ike2', 1990, 1992],
            ['ike2', 2000, 2002],
            ['ike2', 2010, 2012],
            ['ike3', 1990, 1993],
            ['ike3', 2000, 2003],
            ['ike3', 2010, 2013],
            ['ike4', 1990, 1994],
            ['ike4', 2000, 2004],
            ['ike4', 2010, 2014],
        ]
    ]
    assert results[1] == [
        dict(zip(['city', 'year', 'amount'], r))
        for r in
        [
            ['mike0', 2050, 2050],
            ['mike0', 2060, 2060],
            ['mike0', 2070, 2070],
            ['mike1', 2050, 2051],
            ['mike1', 2060, 2061],
            ['mike1', 2070, 2071],
            ['mike2', 2050, 2052],
            ['mike2', 2060, 2062],
            ['mike2', 2070, 2072],
            ['mike3', 2050, 2053],
            ['mike3', 2060, 2063],
            ['mike3', 2070, 2073],
            ['mike4', 2050, 2054],
            ['mike4', 2060, 2064],
            ['mike4', 2070, 2074],
        ]
    ]


def test_concatenate():
    from dataflows import concatenate

    f = Flow(
        [
            {'a': 1, 'b': 2},
            {'a': 2, 'b': 3},
            {'a': 3, 'b': 4},
        ],
        [
            {'c': 4, 'd': 5},
            {'c': 5, 'd': 6},
            {'c': 6, 'd': 7},
        ],
        concatenate({
            'f1': ['a'],
            'f2': ['b', 'c'],
            'f3': ['d']
        })
    )
    results, _, _ = f.results()
    assert results[0] == [
        {'f1': 1, 'f2': 2, 'f3': None},
        {'f1': 2, 'f2': 3, 'f3': None},
        {'f1': 3, 'f2': 4, 'f3': None},
        {'f1': None, 'f2': 4, 'f3': 5},
        {'f1': None, 'f2': 5, 'f3': 6},
        {'f1': None, 'f2': 6, 'f3': 7}
    ]


def test_concatenate_multifield():
    from dataflows import concatenate

    f = Flow(
        [
            {'a': 1, 'b': 2, 'c': None},
            {'a': 2, 'b': None, 'c': 3},
            {'a': 3, 'c': 4},
            {'a': 3, 'b': 6, 'c': 4},
        ],
        concatenate({
            'f1': ['a'],
            'f2': ['b', 'c'],
        })
    )
    results, _, _ = f.results()
    assert results[0] == [
        {'f1': 1, 'f2': 2},
        {'f1': 2, 'f2': 3},
        {'f1': 3, 'f2': 4},
        {'f1': 3, 'f2': 4},
    ]



def test_filter_rows():
    from dataflows import filter_rows

    f = Flow(
        [
            {'a': 1, 'b': 3},
            {'a': 2, 'b': 3},
            {'a': 1, 'b': 4},
            {'a': 2, 'b': 4},
        ],
        filter_rows(equals=[dict(a=1)]),
        filter_rows(not_equals=[dict(b=3)]),
    )
    results, _, _ = f.results()
    assert results[0][0] == dict(a=1, b=4)
    assert len(results[0]) == 1
    assert len(results) == 1


def test_filter_rows_callable():
    from dataflows import filter_rows

    f = Flow(
        [
            {'a': 1, 'b': 3},
            {'a': 2, 'b': 3},
            {'a': 1, 'b': 4},
            {'a': 2, 'b': 4},
        ],
        filter_rows(condition=lambda row: row['a'] > 1 and row['b'] < 4),
    )
    results, _, _ = f.results()
    assert results[0][0] == dict(a=2, b=3)
    assert len(results[0]) == 1
    assert len(results) == 1


def test_sort_rows():
    from dataflows import sort_rows

    f = Flow(
        [
            {'a': 1, 'b': 3},
            {'a': 2, 'b': 3},
            {'a': 3, 'b': 1},
            {'a': 4, 'b': 1},
        ],
        sort_rows(key='{b}{a}'),
    )
    results, _, _ = f.results()
    assert list(results[0]) == [
        {'a': 3, 'b': 1},
        {'a': 4, 'b': 1},
        {'a': 1, 'b': 3},
        {'a': 2, 'b': 3},
    ]


def test_sort_rows_callable():
    from dataflows import sort_rows

    f = Flow(
        [
            {'a': 1, 'b': 3},
            {'a': 2, 'b': 3},
            {'a': 3, 'b': 1},
            {'a': 4, 'b': 1},
        ],
        sort_rows(key=lambda r: '%04d|%04d' % (r['b'], r['a'])),
    )
    results, _, _ = f.results()
    assert list(results[0]) == [
        {'a': 3, 'b': 1},
        {'a': 4, 'b': 1},
        {'a': 1, 'b': 3},
        {'a': 2, 'b': 3},
    ]


def test_sort_reverse_many_rows():
    from dataflows import sort_rows

    f = Flow(
        ({'a': i, 'b': i % 5} for i in range(1000)),
        sort_rows(key='{b}{a}', reverse=True, batch_size=0),
    )
    results, _, _ = f.results()
    results = results[0]
    assert results[0:2] == [{'a': 999, 'b': 4}, {'a': 994, 'b': 4}]
    assert results[998:1000] == [{'a': 5, 'b': 0}, {'a': 0, 'b': 0}]


def test_sort_rows_number():
    from dataflows import sort_rows

    f = Flow(
        [
            {'a': 0.1},
            {'a': -3},
            {'a': -4},
            {'a': 10},
            {'a': 8},
            {'a': 0},
            {'a': -1000000},
            {'a': 1000000},
            {'a': -0.1},
            {'a': -0.2},
            {'a': 0.2},
            {'a': -1000001},
            {'a': 1000001},
            {'a': 6},
            {'a': -10},
            {'a': -0.001},
            {'a': 0.001},
            {'a': 1},
            {'a': -1},
        ],
        sort_rows(key='{a}'),
    )
    results, _, _ = f.results()
    assert list(results[0]) == [
        {'a': -1000001},
        {'a': -1000000},
        {'a': -10},
        {'a': -4},
        {'a': -3},
        {'a': -1},
        {'a': -0.2},
        {'a': -0.1},
        {'a': -0.001},
        {'a': 0},
        {'a': 0.001},
        {'a': 0.1},
        {'a': 0.2},
        {'a': 1},
        {'a': 6},
        {'a': 8},
        {'a': 10},
        {'a': 1000000},
        {'a': 1000001},
    ]


def test_sort_rows_decimal():
    from decimal import Decimal
    from dataflows import sort_rows, load

    f = Flow(
        load('data/numbers.csv', cast_strategy=load.CAST_WITH_SCHEMA),
        sort_rows(key='{a}'),
    )
    results, dp, _ = f.results()
    assert list(results[0]) == [
        {'a': Decimal('-1000')},
        {'a': Decimal('-0.5')},
        {'a': Decimal('-0.4')},
        {'a': Decimal('0')},
        {'a': Decimal('1.1')},
        {'a': Decimal('2')},
        {'a': Decimal('10')},
        {'a': Decimal('1000')}
    ]


def test_sort_rows_datetime():
    import datetime
    from dataflows import sort_rows

    f = Flow(
        [
            {'a': datetime.date(2000, 1, 3)},
            {'a': datetime.date(2010, 1, 2)},
            {'a': datetime.date(2020, 1, 1)},
        ],
        sort_rows(key='{a}'),
    )
    results, _, _ = f.results()
    assert list(results[0]) == [
        {'a': datetime.date(2000, 1, 3)},
        {'a': datetime.date(2010, 1, 2)},
        {'a': datetime.date(2020, 1, 1)},
    ]


def test_deduplicate():
    from dataflows import deduplicate, set_primary_key

    a = [
            {'a': 1, 'b': 3, 'c': 'First'},
            {'a': 2, 'b': 3, 'c': 'First'},
            {'a': 1, 'b': 3, 'c': '!First'},
            {'a': 1, 'b': 2, 'c': 'First'},
            {'a': 2, 'b': 3, 'c': '!First'},
        ]

    f = Flow(
        a,
        set_primary_key(['a', 'b']),
        deduplicate(),
    )
    results, _, _ = f.results()
    assert set(x['c'] for x in results[0]) == {'First'}


def test_duplicate():
    from dataflows import duplicate

    a = [
            {'a': 1, 'b': 3},
            {'a': 2, 'b': 3},
            {'a': 3, 'b': 1},
            {'a': 4, 'b': 1},
        ]

    f = Flow(
        a,
        duplicate(),
    )
    results, _, _ = f.results()
    assert list(results[0]) == a
    assert list(results[1]) == a


def test_duplicate_many_rows():
    from dataflows import duplicate

    f = Flow(
        ({'a': i, 'b': i} for i in range(1000)),
        duplicate(),
    )

    results, _, _ = f.results()
    assert len(results[0]) == 1000
    assert len(results[1]) == 1000

    f = Flow(
        ({'a': i, 'b': i} for i in range(10000)),
        duplicate(batch_size=0),
    )

    results, _, _ = f.results()
    assert len(results[0]) == 10000
    assert len(results[1]) == 10000

def test_duplicate_to_end():
    from dataflows import duplicate

    a = [
        {"a": 1, "b": 3},
    ]
    b = [
        {"c": 1, "d": 3},
    ]

    f = Flow(
        a,
        b,
        duplicate(duplicate_to_end=True),
    )
    results, _, _ = f.results()
    assert list(results[0]) == a
    assert list(results[1]) == b
    assert list(results[2]) == a





def test_flow_as_step():
    def upper(row):
        for k in row:
            row[k] = row[k].upper()

    def lower_first_letter(row):
        for k in row:
            row[k] = row[k][0].lower() + row[k][1:]

    text_processing_flow = Flow(upper, lower_first_letter)

    assert Flow([{'foo': 'bar'}], text_processing_flow).results()[0] == [[{'foo': 'bAR'}]]


def test_load_from_package():
    from dataflows import dump_to_path, load

    Flow(
        [{'foo': 'bar', 'moo': 12}],
        dump_to_path('out/load_from_package')
    ).process()

    ds = Flow(
        load('out/load_from_package/datapackage.json')
    ).datastream()

    assert len(ds.dp.resources) == 1
    assert [list(res) for res in ds.res_iter] == [[{'foo': 'bar', 'moo': 12}]]


def test_load_from_package_zip():
    from dataflows import load

    ds = Flow(
        load('data/load_from_package.zip', format='datapackage')
    ).datastream()

    assert len(ds.dp.resources) == 1
    assert [list(res) for res in ds.res_iter] == [[{'foo': 'bar', 'moo': 12}]]


def test_load_from_env_var():
    import os
    from dataflows import load, dump_to_path

    Flow(
        [{'foo': 'bar'}],
        dump_to_path('out/load_from_env_var')
    ).process()

    os.environ['MY_DATAPACKAGE'] = 'out/load_from_env_var/datapackage.json'
    results, dp, _ = Flow(
        load('env://MY_DATAPACKAGE')
    ).results()

    assert len(dp.resources) == 1
    assert results == [[{'foo': 'bar'}]]


def test_load_from_package_resource_matching():
    from dataflows import dump_to_path, load

    Flow(
        [{'foo': 'bar'}],
        [{'foo': 'baz'}],
        dump_to_path('out/load_from_package_resource_matching(')
    ).process()

    ds = Flow(
        load('out/load_from_package_resource_matching(/datapackage.json', resources=['res_2'])
    ).datastream()

    assert len(ds.dp.resources) == 1
    assert [list(res) for res in ds.res_iter] == [[{'foo': 'baz'}]]


def test_load_strategies():
    from dataflows import load

    i_strategies = [
        load.INFER_FULL,
        load.INFER_PYTHON_TYPES,
        load.INFER_STRINGS,
    ]

    c_strategies = [
        load.CAST_DO_NOTHING,
        load.CAST_TO_STRINGS,
        load.CAST_WITH_SCHEMA
    ]

    res = {}
    for i_s in i_strategies:
        for c_s in c_strategies:
            ret = res.setdefault(i_s + ' ' + c_s, [])
            Flow(
                load('data/beatles_age.json', infer_strategy=i_s, cast_strategy=c_s, on_error=load.ERRORS_DROP),
                load('data/beatles_age.csv', infer_strategy=i_s, cast_strategy=c_s, on_error=load.ERRORS_DROP),
                lambda row: ret.append(row) or row
            ).process()
    out_t = [{'age': 18, 'name': 'john'},
             {'age': 16, 'name': 'paul'},
             {'age': 17, 'name': 'george'},
             {'age': 22, 'name': 'ringo'}]
    out_s = [{'age': '18', 'name': 'john'},
             {'age': '16', 'name': 'paul'},
             {'age': '17', 'name': 'george'},
             {'age': '22', 'name': 'ringo'}]

    assert res == {
        'full nothing': out_t + out_s,
        'full schema': out_t + out_t,
        'full strings': out_s + out_s,
        'pytypes nothing': out_t + out_s,
        'pytypes schema': out_t + out_s,
        'pytypes strings': out_s + out_s,
        'strings nothing': out_s + out_s,
        'strings schema': out_s + out_s,
        'strings strings': out_s + out_s
    }


def test_load_strategy_infer_strings_from_native_types():
    from dataflows import load

    flow = Flow(
        load(
            'data/beatles_age.json',
            infer_strategy='strings',
        ),
    )
    data, package, stats = flow.results()
    assert data == [[
        {'age': '18', 'name': 'john'},
        {'age': '16', 'name': 'paul'},
        {'age': '17', 'name': 'george'},
        {'age': '22', 'name': 'ringo'},
    ]]


def test_load_name_path():
    from dataflows import load

    dp, *_ = Flow(
        load('data/beatles_age.json', name='foo'),
        load('data/beatles_age.csv')
    ).process()

    print(dp.descriptor['resources'])

    res0 = dp.resources[0]
    res1 = dp.resources[1]

    assert res0.name == 'foo'
    assert res0.descriptor['path'] == 'foo.json'
    assert res1.name == 'beatles_age'
    assert res1.descriptor['path'] == 'beatles_age.csv'


def test_load_from_package_resources():
    from dataflows import load

    datapackage = {'resources': [{'name': 'my-resource-{}'.format(i),
                                  'path': 'my-resource-{}.csv'.format(i),
                                  'schema': {'fields': [{'name': 'foo', 'type': 'string'}]}} for i in range(2)]}
    resources = ((row for row in [{'foo': 'bar{}'.format(i)}, {'foo': 'baz{}'.format(i)}]) for i in range(2))

    data, dp, *_ = Flow(
        load((datapackage, resources), resources=['my-resource-1']),
    ).results()

    assert len(dp.resources) == 1
    assert dp.get_resource('my-resource-1').descriptor['path'] == 'my-resource-1.csv'
    assert data[0][1] == {'foo': 'baz1'}


def test_checkpoint():
    from collections import defaultdict
    from dataflows import Flow, checkpoint
    import shutil

    shutil.rmtree('.checkpoints/test_checkpoint', ignore_errors=True)

    stats = defaultdict(int)

    def get_data_count_views():
        stats['stale'] += 1

        def data():
            yield {'foo': 'bar'}
            stats['fresh'] += 1
            stats['stale'] -= 1

        return data()

    def run_data_count_flow():
        assert Flow(
            get_data_count_views(),
            checkpoint('test_checkpoint'),
        ).results()[0] == [[{'foo': 'bar'}]]

    run_data_count_flow()
    run_data_count_flow()
    run_data_count_flow()
    assert stats['fresh'] == 1
    assert stats['stale'] == 2


def test_load_from_checkpoint():
    from dataflows import Flow, checkpoint
    import shutil

    shutil.rmtree('.checkpoints/test_load_from_checkpoint', ignore_errors=True)

    assert Flow(
        [{'foo': 'bar'}],
        checkpoint('test_load_from_checkpoint')
    ).process()

    assert Flow(
        checkpoint('test_load_from_checkpoint')
    ).results()[0] == [[{'foo': 'bar'}]]

def test_sources_iterables():
    from dataflows import Flow, sources

    iterables = [[dict(x=a, y=i) for i in range(3)] for a in ['a', 'b', 'c']]
    res, dp, _ = Flow(
        sources(*iterables)
    ).results()
    assert len(dp.resources)==3
    assert res == iterables

def test_sources_iterables_previous():
    from dataflows import Flow, sources

    previous = [[dict(x=a, y=i) for i in range(3)] for a in ['d', 'e', 'f']]
    iterables = [[dict(x=a, y=i) for i in range(3)] for a in ['a', 'b', 'c']]
    res, dp, _ = Flow(
        *previous,
        sources(*iterables)
    ).results()
    assert len(dp.resources)==6
    assert res == previous + iterables

def test_sources_load():
    from dataflows import Flow, sources, load

    iterables = [[dict(x=a, y=i) for i in range(3)] for a in ['a', 'b', 'c']]
    previous = [[dict(x=a, y=i) for i in range(3)] for a in ['d', 'e', 'f']]
    expected = [
        {'age': '18', 'name': 'john'},
        {'age': '16', 'name': 'paul'},
        {'age': '17', 'name': 'george'},
        {'age': '22', 'name': 'ringo'},
    ]

    flow = Flow(
        sources(
            *previous,
            load(
                'data/beatles_age.json',
                infer_strategy='strings',
            ),
            *iterables
        )
    )
    res, dp, _ = flow.results()
    assert len(dp.resources)==7
    assert res == previous + [expected] + iterables


def test_update_resource():
    from dataflows import Flow, printer, update_resource

    f = Flow(
        *[
            ({k: x} for x in range(10))
            for k in 'abcdef'
        ],
        update_resource(['res_1', 'res_3', 'res_5'], source='thewild'),
        printer()
    )
    results, dp, stats = f.results()
    print(dp.descriptor)
    assert dp.descriptor['resources'][0]['source'] == 'thewild'
    assert dp.descriptor['resources'][2]['source'] == 'thewild'
    assert dp.descriptor['resources'][4]['source'] == 'thewild'


def test_update_schema():
    from dataflows import Flow, printer, update_schema, validate

    f = Flow(
        [['a', '-'], ['a', 0]],
        update_schema(-1, missingValues=['-']),
        validate(),
        printer()
    )
    results, dp, stats = f.results()
    print(dp.descriptor)
    assert results[0] == [
        dict(col0='a', col1=None),
        dict(col0='a', col1=0),
    ]

def test_set_type_resources():
    from dataflows import Flow, set_type, validate

    f = Flow(
        [dict(a=str(i)) for i in range(10)],
        [dict(b=str(i)) for i in range(10)],
        [dict(c='0_' + str(i)) for i in range(10)],
        set_type('a', resources='res_[1]', type='integer'),
        set_type('b', resources=['res_2'], type='integer'),
        set_type('[cd]', resources=-1, type='number', groupChar='_'),
        validate()
    )
    results, dp, stats = f.results()
    print(dp.descriptor)
    assert results[0][1]['a'] == 1
    assert results[1][3]['b'] == 3
    assert results[2][8]['c'] == 8.0


def test_set_type_errors():
    from dataflows import Flow, set_type, ValidationError, exceptions
    from dataflows.base.schema_validator import ignore, drop, raise_exception, clear

    data = [
        {'a': 1, 'b': 1},
        {'a': 2, 'b': 2},
        {'a': 3, 'b': 3},
        {'a': 4, 'b': 'a'},
    ]

    f = Flow(
        data,
        set_type('b', type='integer', on_error=drop),
    )
    results, *_ = f.results()
    assert results[0] == data[:3]

    f = Flow(
        data,
        set_type('b', type='integer', on_error=ignore),
    )
    results, *_ = f.results(on_error=ignore)
    assert results[0] == data[:4]

    f = Flow(
        data,
        set_type('b', type='integer', on_error=clear),
    )
    results, *_ = f.results(on_error=ignore)
    assert results[0][:3] == data[:3]
    assert results[0][3] == dict(a=4, b=None)

    f = Flow(
        data,
        set_type('b', type='integer', on_error=raise_exception),
    )

    with pytest.raises(exceptions.ProcessorError) as excinfo:
        results, *_ = f.results()
    assert isinstance(excinfo.value.cause, ValidationError)

    f = Flow(
        data,
        set_type('b', type='integer'),
    )
    with pytest.raises(exceptions.ProcessorError) as excinfo:
        results, *_ = f.results()
    assert isinstance(excinfo.value.cause, ValidationError)

def test_set_type_transform():
    from dataflows import Flow, set_type, validate, exceptions

    data = [
        dict(a=i) for i in range(5)
    ]

    with pytest.raises(exceptions.ProcessorError):
        Flow(
            data,
            set_type('a', type='integer'),
            validate(),
            set_type('a', type='string'),
            validate(),
        ).process()

    r = Flow(
        data,
        set_type('a', type='integer'),
        validate(),
        set_type('a', type='string', transform=str),
        validate(),
    ).results()[0][0][0]
    assert r['a'] == '0'

    r = Flow(
        data,
        set_type('a', type='integer'),
        validate(),
        set_type('a', type='string', transform=lambda v: str(v)),
        validate(),
    ).results()[0][0][0]
    assert r['a'] == '0'

    r = Flow(
        data,
        set_type('a', type='integer'),
        validate(),
        set_type('a', type='string', transform=lambda v, field_name: field_name),
        validate(),
    ).results()[0][0][0]
    assert r['a'] == 'a'


def test_dump_to_path_use_titles():
    from dataflows import Flow, dump_to_path, set_type
    import tabulator

    Flow(
        [{'hello': 'world', 'hola': 'mundo'}, {'hello': 'עולם', 'hola': 'عالم'}],
        *(set_type(name, resources=['res_1'], title=title) for name, title
          in (('hello', 'שלום'), ('hola', 'aloha'))),
        dump_to_path('out/dump_with_titles', use_titles=True)
    ).process()

    with tabulator.Stream('out/dump_with_titles/res_1.csv') as stream:
        assert stream.read() == [['שלום',   'aloha'],
                                 ['world',  'mundo'],
                                 ['עולם',   'عالم']]


def test_load_dates():
    # from dateutil.tz import tzutc
    from dataflows import Flow, dump_to_path, load, set_type, ValidationError, exceptions
    import datetime

    _today = datetime.date.today()
    _now = datetime.datetime.now()

    def run_flow(datetime_format=None):
        Flow(
            [{'today': str(_today), 'now': str(_now)}],
            set_type('today', type='date'),
            set_type('now', type='datetime', format=datetime_format),
            dump_to_path('out/dump_dates')
        ).process()

    with pytest.raises(exceptions.ProcessorError) as excinfo:
        run_flow()
    assert isinstance(excinfo.value.cause, ValidationError)

    # Default is isoformat(), str() gives a slightly different format:
    # >>> from datetime import datetime
    # >>> n = datetime.now()
    # >>> str(n)
    # '2018-11-22 13:25:47.945209'
    # >>> n.isoformat()
    # '2018-11-22T13:25:47.945209'
    run_flow(datetime_format='%Y-%m-%d %H:%M:%S.%f')

    out_now = datetime.datetime(_now.year, _now.month, _now.day, _now.hour, _now.minute, _now.second)

    assert Flow(
        load('out/dump_dates/datapackage.json'),
    ).results()[0] == [[{'today': _today, 'now': out_now}]]


def test_load_dates_timezones():
    from dataflows import Flow, checkpoint
    from datetime import datetime, timezone
    import shutil

    dates = [
        datetime.now().replace(microsecond=0),
        datetime.now(timezone.utc).replace(microsecond=0).astimezone()
    ]

    shutil.rmtree('.checkpoints/test_load_dates_timezones', ignore_errors=True)

    Flow(
        [{'date': d.date(), 'datetime': d} for d in dates],
        checkpoint('test_load_dates_timezones')
    ).process()

    results = Flow(
        checkpoint('test_load_dates_timezones')
    ).results()

    assert list(map(lambda x: x['date'], results[0][0])) == \
        list(map(lambda x: x.date(), dates))
    assert list(map(lambda x: x['datetime'], results[0][0])) == \
        list(map(lambda x: x, dates))


def test_add_field():
    from dataflows import Flow, add_field
    f = Flow(
        (dict(a=i) for i in range(3)),
        add_field('b', 'string', 'b'),
        add_field('c', 'number'),
        add_field('d', 'boolean', title='mybool'),
    )
    results, dp, _ = f.results()
    assert results == [[
        {'a': 0, 'b': 'b', 'c': None, 'd': None},
        {'a': 1, 'b': 'b', 'c': None, 'd': None},
        {'a': 2, 'b': 'b', 'c': None, 'd': None}
    ]]
    assert dp.descriptor == \
        {
            'profile': 'data-package',
            'resources': [
                {
                    'name': 'res_1',
                    'path': 'res_1.csv',
                    'profile': 'tabular-data-resource',
                    'schema': {
                        'fields': [
                            {
                                'format': 'default',
                                'name': 'a',
                                'type': 'integer'
                            },
                            {
                                'format': 'default',
                                'name': 'b',
                                'type': 'string'
                            },
                            {
                                'format': 'default',
                                'name': 'c',
                                'type': 'number'
                            },
                            {
                                'format': 'default',
                                'name': 'd',
                                'title': 'mybool',
                                'type': 'boolean'
                            }
                        ],
                        'missingValues': ['']
                    }
                }
            ]
        }


def test_load_empty_headers():
    from dataflows import Flow, load

    def ensure_type(t):
        def func(row):
            assert isinstance(row['a'], t)
        return func

    results, dp, stats = Flow(load('data/empty_headers.csv'),
                              ensure_type(str)).results()
    assert results[0] == [
        {'a': 1, 'b': 2},
        {'a': 2, 'b': 3},
        {'a': 3, 'b': 4},
        {'a': 5, 'b': 6}
    ]
    assert len(dp.resources[0].schema.fields) == 2

    results, dp, stats = Flow(load('data/empty_headers.csv', validate=True),
                              ensure_type(int)).results()
    assert results[0] == [
        {'a': 1, 'b': 2},
        {'a': 2, 'b': 3},
        {'a': 3, 'b': 4},
        {'a': 5, 'b': 6}
    ]

    results, dp, stats = Flow(load('data/empty_headers.csv', force_strings=True),
                              ensure_type(str)).results()
    assert results[0] == [
        {'a': '1', 'b': '2'},
        {'a': '2', 'b': '3'},
        {'a': '3', 'b': '4'},
        {'a': '5', 'b': '6'}
    ]
    assert len(dp.resources[0].schema.fields) == 2

    results, dp, stats = Flow(load('data/empty_headers.csv', force_strings=True, validate=True),
                              ensure_type(str)).results()
    assert results[0] == [
        {'a': '1', 'b': '2'},
        {'a': '2', 'b': '3'},
        {'a': '3', 'b': '4'},
        {'a': '5', 'b': '6'}
    ]
    assert len(dp.resources[0].schema.fields) == 2


def test_load_xml():
    from dataflows import Flow, load

    results, dp, stats = Flow(load('data/sample.xml')).results()

    assert results[0] == [
        {'publication-year': 1954, 'title': 'The Fellowship of the Ring'},
        {'publication-year': 1954, 'title': 'The Two Towers'},
        {'publication-year': 1955, 'title': 'The Return of the King'}
    ]


def test_save_load_dates():
    from dataflows import Flow, dump_to_path, load, set_type, printer
    import datetime

    Flow(
        [{'id': 1, 'ts': datetime.datetime.now()},
         {'id': 2, 'ts': datetime.datetime.now()}],
        set_type('ts', type='datetime', format='%Y-%m-%d/%H:%M:%S'),
        dump_to_path('out/test_save_load_dates')
    ).process()

    res, _, _ = Flow(
        load('out/test_save_load_dates/datapackage.json'),
        printer()
    ).results()


def test_stream_simple():
    from dataflows import stream, unstream

    datas1 = [
        {'a': 1, 'b': True, 'c': 'c1'},
        {'a': 2, 'b': True, 'c': 'c2'},
    ]
    datas2 = [
        {'a': 3, 'b': True, 'c': 'c3'},
        {'a': 4, 'b': True, 'c': 'c4'},
        {'a': 5, 'b': True, 'c': 'c5'},
    ]
    Flow(
        datas1,
        datas2,
        stream(open('out/test_stream_simple.stream', 'w'))
    ).process()

    results, dp, _ = Flow(
        unstream(open('out/test_stream_simple.stream'))
    ).results()

    assert results[0] == datas1
    assert results[1] == datas2


def test_stream_bad_dates():
    from dataflows import stream, unstream, set_type, dump_to_path
    import datetime

    datas1 = [
        {'a': '0001/1/1'},
    ]
    Flow(
        datas1,
        set_type('a', type='date', format='%Y/%m/%d'),
        stream(open('out/test_stream_bad_dates.stream', 'w'))
    ).process()

    results, dp, _ = Flow(
        unstream(open('out/test_stream_bad_dates.stream')),
        dump_to_path('out/test_stream_bad_dates')
    ).results()

    assert results[0][0]['a'] == datetime.date(1,1,1)



def test_set_primary_key():
    from dataflows import set_primary_key

    datas1 = [
        {'a': 1, 'b': True, 'c': 'c1'},
        {'a': 2, 'b': True, 'c': 'c2'},
    ]
    _, dp, _ = Flow(
        datas1,
        set_primary_key(['a', 'b']),
    ).results()

    assert dp.resources[0].schema.primary_key == ['a', 'b']


def test_validate():
    from dataflows import Flow, validate, set_type
    from dataflows.base.schema_validator import ignore
    data = [
        {'a': 1, 'b': 1},
        {'a': 2, 'b': 2},
        {'a': 3, 'b': 3},
        {'a': 4, 'b': 'a'},
    ]

    class on_error():
        def __init__(self):
            self.bad_row, self.bad_index = None, None

        def __call__(self, name, row, i, e):
            self.bad_row, self.bad_index = row, i
            return False

    class on_error_fields():
        def __init__(self):
            self.bad_row, self.bad_index, self.bad_field = None, None, None

        def __call__(self, name, row, i, e, field):
            self.bad_row, self.bad_index = row, i
            self.bad_field = field
            return False

    # Schema validator
    handler = on_error()
    res, *_ = Flow(
        data,
        set_type('b', type='integer', on_error=ignore),
        validate(on_error=handler)
    ).results()
    assert len(res[0]) == 3
    assert handler.bad_row == {'a': 4, 'b': 'a'}
    assert handler.bad_index == 3

    # Schema validator with fields
    handler = on_error_fields()
    res, *_ = Flow(
        data,
        set_type('b', type='integer', on_error=ignore),
        validate(on_error=handler)
    ).results()
    assert len(res[0]) == 3
    assert handler.bad_row == {'a': 4, 'b': 'a'}
    assert handler.bad_index == 3
    assert handler.bad_field.name == 'b'

    # Field validator
    handler = on_error()
    res, *_ = Flow(
        data,
        set_type('b', type='integer', on_error=ignore),
        validate('a', lambda v: v < 4, on_error=handler)
    ).results()
    assert len(res[0]) == 3
    assert handler.bad_row == {'a': 4, 'b': 'a'}
    assert handler.bad_index == 3

    # Row validator
    handler = on_error()
    res, *_ = Flow(
        data,
        set_type('b', type='integer', on_error=ignore),
        validate(lambda v: v['a'] < 4, on_error=handler)
    ).results()
    assert len(res[0]) == 3
    assert handler.bad_row == {'a': 4, 'b': 'a'}
    assert handler.bad_index == 3


def test_join():
    from dataflows import Flow, join, join_with_self, set_type, sort_rows
    from decimal import Decimal

    characters = [
        {'first_name': 'Jaime', 'house': 'Lannister', 'last_name': 'Lannister', 'age': 34},
        {'first_name': 'Tyrion', 'house': 'Lannister', 'last_name': 'Lannister', 'age': 27},
        {'first_name': 'Cersei', 'house': 'Lannister', 'last_name': 'Lannister', 'age': 34},
        {'first_name': 'Jon', 'house': 'Stark', 'last_name': 'Snow', 'age': 17},
        {'first_name': 'Sansa', 'house': 'Stark', 'last_name': 'Stark', 'age': 14},
        {'first_name': 'Rickon', 'house': 'Stark', 'last_name': 'Stark', 'age': 5},
        {'first_name': 'Arya', 'house': 'Stark', 'last_name': 'Stark', 'age': 11},
        {'first_name': 'Bran', 'house': 'Stark', 'last_name': 'Stark', 'age': 10},
        {'first_name': 'Daenerys', 'house': 'Targaryen', 'last_name': 'Targaryen', 'age': 16},
    ]

    houses = [
        {'house': 'House of Lannister'},
        {'house': 'House of Greyjoy'},
        {'house': 'House of Stark'},
        {'house': 'House of Targaryen'},
        {'house': 'House of Martell'},
        {'house': 'House of Tyrell'},
    ]

    res, _, _ = Flow(
        characters,
        set_type('age', type='number'),
        houses,
        join(
            'res_1',
            'House of {house}',
            'res_2',
            '{house}',
            dict(
                max_age={
                    'name': 'age',
                    'aggregate': 'max'
                },
                avg_age={
                    'name': 'age',
                    'aggregate': 'avg'
                },
                representative={
                    'name': 'first_name',
                    'aggregate': 'last'
                },
                representative_age={
                    'name': 'age'
                },
                number_of_characters={
                    'aggregate': 'count'
                },
                last_names={
                    'name': 'last_name',
                    'aggregate': 'counters'
                }
            ), full=False, source_delete=True
        )
    ).results()

    assert res[0] == [
        {
            'avg_age': Decimal('31.66666666666666666666666667'),
            'house': 'House of Lannister',
            'max_age': Decimal(34),
            'number_of_characters': 3,
            'representative': 'Cersei',
            'representative_age': Decimal(34),
            'last_names': [('Lannister', 3)]
        },
        {
            'avg_age': Decimal('11.4'),
            'house': 'House of Stark',
            'max_age': Decimal(17),
            'number_of_characters': 5,
            'representative': 'Bran',
            'representative_age': Decimal(10),
            'last_names': [('Stark', 4), ('Snow', 1)]
        },
        {
            'avg_age': Decimal(16),
            'house': 'House of Targaryen',
            'max_age': Decimal(16),
            'number_of_characters': 1,
            'representative': 'Daenerys',
            'representative_age': Decimal(16),
            'last_names': [('Targaryen', 1)]
        },
    ]

    # Find youngest of each house
    res, _, _ = Flow(
        characters,
        set_type('age', type='number'),
        sort_rows('{age:02}'),
        join_with_self(
            'res_1',
            '{house}',
            {
                'the_house': {
                    'name': 'house'
                },
                '*': {
                    'aggregate': 'first'
                },
            }
        ),
        sort_rows('{the_house}')
    ).results()

    assert res[0] == [
        {
            'the_house': 'Lannister',
            'first_name': 'Tyrion',
            'last_name': 'Lannister',
            'age': Decimal('27')
        },
        {
            'the_house': 'Stark',
            'first_name': 'Rickon',
            'last_name': 'Stark',
            'age': Decimal('5')
        },
        {
            'the_house': 'Targaryen',
            'first_name': 'Daenerys',
            'last_name': 'Targaryen',
            'age': Decimal('16')
        }
    ]


def test_load_limit_rows():
    from dataflows import load
    flow = Flow(
        load('data/beatles.csv', limit_rows=3)
    )
    data = flow.results()[0]
    assert data == [[
        {'name': 'john', 'instrument': 'guitar'},
        {'name': 'paul', 'instrument': 'bass'},
        {'name': 'george', 'instrument': 'guitar'},
    ]]


def test_set_type_regex():
    from dataflows import load, set_type
    flow = Flow(
        load('data/regex.csv'),
        set_type('city', type='string'),
        set_type('temperature (24h)', type='integer', regex=False),
    )
    data = flow.results()[0]
    assert data == [[
        {'city': 'london', 'temperature (24h)': 23},
        {'city': 'paris', 'temperature (24h)': 26},
        {'city': 'rome', 'temperature (24h)': 21},
    ]]


def test_load_override_schema():
    from dataflows import load
    flow = Flow(
        load('data/beatles_age.csv',
            override_schema={
                'title': 'title',
                'missingValues': ['ringo'],
            }
        ),
    )
    data, package, stats = flow.results()
    assert package.descriptor == {
        'profile': 'data-package',
        'resources': [{
            'format': 'csv',
            'name': 'beatles_age',
            'path': 'beatles_age.csv',
            'profile': 'tabular-data-resource',
            'schema': {
                'fields': [
                    {'format': 'default', 'name': 'name', 'type': 'string'},
                    {'format': 'default', 'name': 'age', 'type': 'integer'}
                ],
                'missingValues': ['ringo'],
                'title': 'title'
            }
        }]
    }
    assert data == [[
        {'name': 'john', 'age': 18},
        {'name': 'paul', 'age': 16},
        {'name': 'george', 'age': 17},
        {'name': None, 'age': 22},
    ]]


def test_load_override_schema_and_fields():
    from dataflows import load
    flow = Flow(
        load('data/beatles_age.csv',
            override_schema={
                'title': 'title',
                'missingValues': ['ringo'],
            },
            override_fields={
                'age': {'type': 'string'},
            }
        ),
    )
    data, package, stats = flow.results()
    assert package.descriptor == {
        'profile': 'data-package',
        'resources': [{
            'format': 'csv',
            'name': 'beatles_age',
            'path': 'beatles_age.csv',
            'profile': 'tabular-data-resource',
            'schema': {
                'fields': [
                    {'format': 'default', 'name': 'name', 'type': 'string'},
                    {'format': 'default', 'name': 'age', 'type': 'string'}
                ],
                'missingValues': ['ringo'],
                'title': 'title',
            }
        }]
    }
    assert data == [[
        {'name': 'john', 'age': '18'},
        {'name': 'paul', 'age': '16'},
        {'name': 'george', 'age': '17'},
        {'name': None, 'age': '22'},
    ]]


def test_delete_fields_regex():
    from dataflows import load, delete_fields
    flow = Flow(
        load('data/regex.csv'),
        delete_fields(['temperature (24h)'], regex=False),
    )
    data = flow.results()[0]
    assert data == [[
        {'city': 'london'},
        {'city': 'paris'},
        {'city': 'rome'},
    ]]


def test_join_full_outer():
    from dataflows import load, set_type, join
    flow = Flow(
        load('data/population.csv'),
        load('data/cities.csv'),
        join(
            source_name='population',
            source_key=['id'],
            target_name='cities',
            target_key=['id'],
            fields={'population': {'name': 'population'}},
            mode='full-outer',
        ),
    )
    data = flow.results()[0]
    assert data == [[
        {'id': 1, 'city': 'london', 'population': 8},
        {'id': 2, 'city': 'paris', 'population': 2},
        {'id': 3, 'city': 'rome', 'population': None},
        {'id': 4, 'city': None, 'population': 3},
    ]]


def test_join_row_number():
    from dataflows import load, set_type, join
    flow = Flow(
        load('data/population.csv'),
        load('data/cities.csv'),
        join(
            source_name='population',
            source_key=['#'],
            target_name='cities',
            target_key=['#'],
            fields={'population': {'name': 'population'}}
        ),
    )
    data = flow.results()[0]
    assert data == [[
        {'id': 1, 'city': 'london', 'population': 8},
        {'id': 2, 'city': 'paris', 'population': 2},
        {'id': 3, 'city': 'rome', 'population': 3},
    ]]


def test_join_row_number_readme_example():
    from dataflows import load, set_type, join
    flow = Flow(
        load('data/values.csv'),
        load('data/names.csv'),
        join(
            source_name='values',
            source_key=['#'],
            target_name='names',
            target_key=['#'],
            fields={'values': {'name': 'values'}}
        ),
    )
    data = flow.results()[0]
    assert data == [[
        {'id': 1, 'names': 'name1', 'values': 'value1'},
        {'id': 2, 'names': 'name2', 'values': 'value2'},
    ]]


def test_join_row_number_format_string():
    from dataflows import load, set_type, join
    flow = Flow(
        load('data/population.csv'),
        load('data/cities_comment.csv'),
        join(
            source_name='population',
            source_key='city with population in row {#}',
            target_name='cities_comment',
            target_key='{comment}',
            fields={'population': {'name': 'population'}}
        ),
    )
    data = flow.results()[0]
    assert data == [[
        {'city': 'paris', 'population': 2, 'comment': 'city with population in row 2'},
        {'city': 'london', 'population': 8, 'comment': 'city with population in row 1'},
        {'city': 'rome', 'population': 3, 'comment': 'city with population in row 3'},
    ]]


def test_join_preserve_source_fields_order():
    from dataflows import load, join
    flow = Flow(
        load('data/cities_metadata.csv'),
        load('data/cities.csv'),
        join(
            source_name='cities_metadata',
            source_key='{id}',
            target_name='cities',
            target_key='{id}',
            fields={'key1': {'name': 'key1'}, 'key2': {'name': 'key2'}}
        ),
    )
    data, package, stats = flow.results()
    assert package.descriptor['resources'][0]['schema']['fields'] == [
        {'name': 'id', 'type': 'integer', 'format': 'default'},
        {'name': 'city', 'type': 'string', 'format': 'default'},
        {'name': 'key2', 'type': 'string', 'format': 'default'},
        {'name': 'key1', 'type': 'string', 'format': 'default'}
    ]
    assert data == [[
        {'id': 1, 'city': 'london', 'key1': 'val1', 'key2': 'val2'},
        {'id': 2, 'city': 'paris', 'key1': 'val1', 'key2': 'val2'},
        {'id': 3, 'city': 'rome', 'key1': 'val1', 'key2': 'val2'},
    ]]


def test_load_duplicate_headers():
    from dataflows import load, exceptions
    flow = Flow(
        load('data/duplicate_headers.csv'),
    )
    with pytest.raises(exceptions.ProcessorError) as excinfo:
        flow.results()
    cause = excinfo.value.cause
    assert 'duplicate headers' in str(cause)


def test_load_duplicate_headers_with_deduplicate_headers_flag():
    from dataflows import load
    flow = Flow(
        load('data/duplicate_headers.csv', deduplicate_headers=True),
    )
    data, package, stats = flow.results()
    assert package.descriptor['resources'][0]['schema']['fields'] == [
        {'name': 'header1', 'type': 'string', 'format': 'default'},
        {'name': 'header2 (1)', 'type': 'string', 'format': 'default'},
        {'name': 'header2 (2)', 'type': 'string', 'format': 'default'},
    ]
    assert data == [[
        {'header1': 'value1', 'header2 (1)': 'value2', 'header2 (2)': 'value3'},
    ]]


# Temporal format

def test_force_temporal_format():
    import datetime
    from dataflows import load, update_resource, dump_to_path

    # Dump
    Flow(
        load('data/temporal.csv',
            name='temporal',
            override_fields={
                'datetime': {'type': 'datetime', 'outputFormat': '%y|%m|%d %H|%M|%S'},
                'date': {'outputFormat': '%y|%m|%d'},
                'time': {'outputFormat': '%H|%M|%S'},
            }),
        dump_to_path('out/force_temporal_format',
            temporal_format_property='outputFormat')
    ).process()

    # Load
    flow = Flow(
        load('out/force_temporal_format/datapackage.json')
    )
    data, package, stats = flow.results()

    # Assert
    assert package.descriptor['resources'][0]['schema'] == {
        'fields': [
            {'format': 'default', 'name': 'event', 'type': 'string'},
            {'format': '%y|%m|%d %H|%M|%S', 'name': 'datetime', 'type': 'datetime'},
            {'format': '%y|%m|%d', 'name': 'date', 'type': 'date'},
            {'format': '%H|%M|%S', 'name': 'time', 'type': 'time'},
        ],
        'missingValues': [''],
    }
    assert data == [[
        {
            'event': 'start',
            'datetime': datetime.datetime(2015, 1, 2, 15, 30, 45),
            'date': datetime.date(2015, 1, 2),
            'time': datetime.time(15, 30, 45),
        },
        {
            'event': 'end',
            'datetime': datetime.datetime(2016, 6, 25, 8, 10, 4),
            'date': datetime.date(2016, 6, 25),
            'time': datetime.time(8, 10, 4),
        }
    ]]


# Extract missing values

def test_extract_missing_values():
    from dataflows import load
    schema = {
        'missingValues': ['err1', 'err2', 'mis1', 'mis2'],
        'fields': [
            {'name': 'col1', 'type': 'number', 'format': 'default'},
            {'name': 'col2', 'type': 'number', 'format': 'default'},
        ]
    }
    flow = Flow(
        load('data/missing_values.csv', override_schema=schema, extract_missing_values=True),
    )
    data, package, stats = flow.results()
    assert package.descriptor['resources'][0]['schema']['fields'][0] == schema['fields'][0]
    assert package.descriptor['resources'][0]['schema']['fields'][1] == schema['fields'][1]
    assert package.descriptor['resources'][0]['schema']['fields'][2] == {
        'name': 'missingValues',
        'type': 'object',
        'format': 'default',
        'values': schema['missingValues'],
    }
    assert data == [[
        {'col1': 1, 'col2': 1, 'missingValues': {}},
        {'col1': None, 'col2': 2, 'missingValues': {'col1': 'err1'}},
        {'col1': 3, 'col2': 3, 'missingValues': {}},
        {'col1': 4, 'col2': None, 'missingValues': {'col2': 'err2'}},
        {'col1': 5, 'col2': 5, 'missingValues': {}},
        {'col1': None, 'col2': None, 'missingValues': {'col1': 'mis1', 'col2': 'mis2'}},
        {'col1': 7, 'col2': 7, 'missingValues': {}},
    ]]

def test_extract_missing_values_options():
    from dataflows import load
    schema = {
        'missingValues': ['err1', 'err2', 'mis1', 'mis2'],
        'fields': [
            {'name': 'col1', 'type': 'number', 'format': 'default'},
            {'name': 'col2', 'type': 'number', 'format': 'default'},
        ]
    }
    flow = Flow(
        load('data/missing_values.csv', override_schema=schema, extract_missing_values={
            'source': 'col1',
            'target': 'notes'
        }),
    )
    data, package, stats = flow.results()
    assert package.descriptor['resources'][0]['schema']['fields'][0] == schema['fields'][0]
    assert package.descriptor['resources'][0]['schema']['fields'][1] == schema['fields'][1]
    assert package.descriptor['resources'][0]['schema']['fields'][2] == {
        'name': 'notes',
        'type': 'object',
        'format': 'default',
        'values': schema['missingValues'],
    }
    assert data == [[
        {'col1': 1, 'col2': 1, 'notes': {}},
        {'col1': None, 'col2': 2, 'notes': {'col1': 'err1'}},
        {'col1': 3, 'col2': 3, 'notes': {}},
        {'col1': 4, 'col2': None, 'notes': {}},
        {'col1': 5, 'col2': 5, 'notes': {}},
        {'col1': None, 'col2': None, 'notes': {'col1': 'mis1'}},
        {'col1': 7, 'col2': 7, 'notes': {}},
    ]]


def test_extract_missing_values_options_source_is_list():
    from dataflows import load
    schema = {
        'missingValues': ['err1', 'err2', 'mis1', 'mis2'],
        'fields': [
            {'name': 'col1', 'type': 'number', 'format': 'default'},
            {'name': 'col2', 'type': 'number', 'format': 'default'},
        ]
    }
    flow = Flow(
        load('data/missing_values.csv', override_schema=schema, extract_missing_values={
            'source': ['col1', 'col2'],
        }),
    )
    data, package, stats = flow.results()
    assert package.descriptor['resources'][0]['schema']['fields'][0] == schema['fields'][0]
    assert package.descriptor['resources'][0]['schema']['fields'][1] == schema['fields'][1]
    assert package.descriptor['resources'][0]['schema']['fields'][2] == {
        'name': 'missingValues',
        'type': 'object',
        'format': 'default',
        'values': schema['missingValues'],
    }
    assert data == [[
        {'col1': 1, 'col2': 1, 'missingValues': {}},
        {'col1': None, 'col2': 2, 'missingValues': {'col1': 'err1'}},
        {'col1': 3, 'col2': 3, 'missingValues': {}},
        {'col1': 4, 'col2': None, 'missingValues': {'col2': 'err2'}},
        {'col1': 5, 'col2': 5, 'missingValues': {}},
        {'col1': None, 'col2': None, 'missingValues': {'col1': 'mis1', 'col2': 'mis2'}},
        {'col1': 7, 'col2': 7, 'missingValues': {}},
    ]]


def test_conditional():
    from dataflows import Flow, conditional, add_field

    tester = lambda dp: 'b' in [f.name for r in dp.resources for f in r.schema.fields]

    data1 = [
        dict(a=i, b=i) for i in range(3)
    ]
    data2 = [
        dict(a=i, c=i) for i in range(3)
    ]

    result1, _, _ = Flow(
        data1, conditional(tester, Flow(add_field('d', 'integer', lambda r: r['a'])))
    ).results()
    result2, _, _ = Flow(
        data2, conditional(tester, Flow(add_field('d', 'integer', lambda r: r['a'])))
    ).results()
    def duplicate_last_field(dp):
        last_field = dp.descriptor['resources'][0]['schema']['fields'][-1]['name']
        return Flow(add_field(last_field + '_1', 'integer', lambda r: r['a']))
    result3, _, _ = Flow(
        data1, conditional(tester, duplicate_last_field)
    ).results()
    assert result1[0] == [
        dict(a=i, b=i, d=i) for i in range(3)
    ]
    assert result2[0] == [
        dict(a=i, c=i) for i in range(3)
    ]
    assert result3[0] == [
        dict(a=i, b=i, b_1=i) for i in range(3)
    ]


def test_finalizer():
    from dataflows import Flow, finalizer

    stats = dict(
        processed=0,
        detected=None
    )

    def process(row):
        stats['processed'] += 1

    def finalize():
        stats['detected'] = stats['processed']

    Flow(
        (dict(a=1) for i in range(10)),
        process,
        finalizer(finalize),
    ).process()

    assert stats['processed'] == 10
    assert stats['detected'] == 10


def test_finalizer_with_stats():
    from dataflows import Flow, finalizer, update_stats

    visited = dict(visited=False)

    def finalize(stats={}):
        assert stats['processed'] == 100
        assert stats['detected'] == 200
        visited['visited'] = True

    Flow(
        (dict(a=1) for i in range(10)),
        update_stats(dict(processed=100)),
        update_stats(dict(detected=200)),
        finalizer(finalize),
    ).process()

    assert visited['visited']

def test_finalizer_not_last():
    from dataflows import Flow, finalizer, update_stats, printer

    visited = dict(visited=False)

    def finalize(stats={}):
        assert stats['processed'] == 100
        assert stats['detected'] == 200
        visited['visited'] = True

    Flow(
        (dict(a=1) for i in range(10)),
        update_stats(dict(processed=100)),
        update_stats(dict(detected=200)),
        finalizer(finalize),
        printer()
    ).process()

    assert visited['visited']

def test_update_stats():

    stats1 =  dict(a=1)
    stats2 =  dict(b=1)
    data = [{'c': 1}, {'c': 2}]

    from dataflows import update_stats, Flow

    res, _, stats = Flow(
        data,
        update_stats(stats1),
        update_stats(stats2),
    ).results()

    assert res == [data]
    assert stats == {'a': 1, 'b': 1}


def test_dump_to_zip():
    from dataflows import Flow, dump_to_zip

    zz = dump_to_zip('out/test_dump_to_zip.zip')
    Flow([dict(a=1)], zz).process()
    assert zz.out_file.closed

def test_dump_to_geojson():
    from dataflows import Flow, dump_to_path, load, add_computed_field, delete_fields
    Flow(
        load('data/cities_location.csv'),
        add_computed_field(target=dict(name='Location', type='geopoint'), operation='format', with_='{lat}, {long}'),
        delete_fields(['lat', 'long']),
        dump_to_path(out_path='out', format='geojson'),
    ).process()


def test_dump_to_excel():
    import datetime
    import openpyxl
    import os
    from dataflows import Flow, dump_to_path, update_resource
    data1 = [
        dict(
            a=str(i) + 'str',
            b=i,
            c=datetime.datetime.now() + datetime.timedelta(days=i),
            d=datetime.date.today() + datetime.timedelta(days=i),
            e=bool(i % 2),
        ) for i in range(10)
    ]
    data2 = [
        dict(
            a=str(i) + 'str',
            b=i,
            c=datetime.datetime.now() + datetime.timedelta(days=i),
            d=datetime.date.today() + datetime.timedelta(days=i),
            e=bool(i % 2),
        ) for i in range(10, 20)
    ]
    filename = 'test_excel/test_excel.xlsx'
    Flow(
        data1,
        update_resource(-1, name='test_update', path='test_excel.xslx'),
        dump_to_path(out_path='test_excel', format='excel', options=dict(sheetname='test1', update_existing=filename)),
    ).process()
    Flow(
        data2,
        update_resource(-1, name='test_update', path='test_excel.xslx'),
        dump_to_path(out_path='test_excel', format='excel', options=dict(sheetname='test2', update_existing=filename)),
    ).process()
    wb = openpyxl.load_workbook(filename)
    assert len(wb.sheetnames) == 2
    assert wb.sheetnames[0] == 'test1'
    assert wb.sheetnames[1] == 'test2'
    assert wb['test2']['A11'].value == '19str'

    Flow(
        data1,
        update_resource(-1, name='test_overwrite', path='test_excel.xlsx'),
        dump_to_path(out_path='test_excel', format='excel'),
    ).process()
    wb = openpyxl.load_workbook(filename)
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == 'test_overwrite'
    assert wb['test_overwrite']['A11'].value == '9str'

    os.unlink(filename)


def test_rename_fields_simple():
    from dataflows import Flow, rename_fields

    data = [dict(a=i, b=i, c=i) for i in range(5)]
    res = Flow(
        data,
        rename_fields(dict(
            a='A', b='B'
        ), regex=False),
    ).results()[0][0]

    assert res == [dict(A=i, B=i, c=i) for i in range(5)]

def test_rename_fields_regex():
    from dataflows import Flow, rename_fields

    data = [dict(a1=i, a2=i, c=i) for i in range(5)]
    res = Flow(
        data,
        rename_fields({
            r'a(\d)': r'A\1'
        }),
    ).results()[0][0]

    assert res == [dict(A1=i, A2=i, c=i) for i in range(5)]

def test_rename_fields_double_rename():
    from dataflows import Flow, rename_fields, exceptions

    data = [dict(a1=i, a2=i, c=i) for i in range(5)]
    with pytest.raises(exceptions.ProcessorError):
        Flow(
            data,
            rename_fields({
                r'a(\d)': r'A'
            }),
        ).results()[0][0]

def test_rename_fields_double_rename_different_resources():
    from dataflows import Flow, rename_fields, exceptions

    data1 = [dict(a1=i, b=i, c=i) for i in range(5)]
    data2 = [dict(a2=i, b=i, c=i) for i in range(5)]
    res = Flow(
        data1,
        data2,
        rename_fields({
            r'a(\d)': r'A'
        }),
    ).results()[0]
    assert res == [
        [dict(A=i, b=i, c=i) for i in range(5)],
        [dict(A=i, b=i, c=i) for i in range(5)],
    ]

def test_rename_fields_specify_resource():
    from dataflows import Flow, rename_fields, exceptions

    data1 = [dict(a1=i, b=i, c=i) for i in range(5)]
    data2 = [dict(a2=i, b=i, c=i) for i in range(5)]
    res = Flow(
        data1,
        data2,
        rename_fields({
            r'a(\d)': r'A'
        }, resources=-1),
    ).results()[0]
    assert res == [
        [dict(a1=i, b=i, c=i) for i in range(5)],
        [dict(A=i, b=i, c=i) for i in range(5)],
    ]

def test_parallelize():
    from dataflows import Flow, parallelize, add_field
    data = [dict(a=i, b=i) for i in range(3300)]
    print('created data')
    def mult(row):
        row['c'] = row['a'] * row['b']

    res = Flow(
        data,
        add_field('c', 'integer'),
        parallelize(mult),
    ).results()[0][0][:100]
    print(res)